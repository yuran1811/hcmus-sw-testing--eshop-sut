#!/usr/bin/env python3
"""
Initialize (or extend) the standard test-case Excel file for HW06.

Usage:
    python3 init_test_case_excel.py --api "FR-02 Login" --out testcases_api1.xlsx
    python3 init_test_case_excel.py --api "FR-08 Checkout" --out testcases_api2.xlsx --append

Requires: pip install openpyxl --break-system-packages
"""
import argparse
import os
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

COLUMNS = [
    "Test_ID", "API", "FR", "Category", "SEC_Ref", "Preconditions",
    "Steps", "Input", "Expected_Result", "Priority", "Source",
    "Audit_Label", "Audit_Reason", "Execution_Status", "Notes",
]

CATEGORY_OPTIONS = '"DomainPartition,StateTransition,Security,Schema"'
SOURCE_OPTIONS = '"AI,Human"'
AUDIT_OPTIONS = '"VALID,INVALID,INCOMPLETE"'
EXEC_OPTIONS = '"Pass,Fail,Blocked,NotRun"'
PRIORITY_OPTIONS = '"High,Medium,Low"'


def build_sheet(ws, api_name):
    ws.title = "TestCases"
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = 18
    ws.column_dimensions["G"].width = 40  # Steps
    ws.column_dimensions["H"].width = 30  # Input
    ws.column_dimensions["I"].width = 40  # Expected_Result
    ws.column_dimensions["M"].width = 35  # Audit_Reason
    ws.freeze_panes = "A2"

    max_rows = 500  # enough room for >=35 test cases/API plus extensions

    dv_category = DataValidation(type="list", formula1=CATEGORY_OPTIONS, allow_blank=True)
    dv_source = DataValidation(type="list", formula1=SOURCE_OPTIONS, allow_blank=True)
    dv_audit = DataValidation(type="list", formula1=AUDIT_OPTIONS, allow_blank=True)
    dv_exec = DataValidation(type="list", formula1=EXEC_OPTIONS, allow_blank=True)
    dv_priority = DataValidation(type="list", formula1=PRIORITY_OPTIONS, allow_blank=True)

    for dv, col_letter in [
        (dv_category, "D"), (dv_source, "K"), (dv_audit, "L"),
        (dv_exec, "N"), (dv_priority, "J"),
    ]:
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}{max_rows}")

    # Prefill the API column with the API name for convenience
    for row in range(2, max_rows + 1):
        ws.cell(row=row, column=2, value=api_name)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--api", required=True, help="API name, e.g. 'FR-02 Login'")
    parser.add_argument("--out", required=True, help="Output .xlsx file path")
    parser.add_argument("--append", action="store_true", help="If the file already exists, add a new sheet instead of overwriting")
    args = parser.parse_args()

    if args.append and os.path.exists(args.out):
        wb = load_workbook(args.out)
        ws = wb.create_sheet(title=f"TestCases_{args.api[:20]}")
        build_sheet(ws, args.api)
    else:
        wb = Workbook()
        ws = wb.active
        build_sheet(ws, args.api)

    wb.save(args.out)
    print(f"Created/updated: {args.out}")


if __name__ == "__main__":
    main()
